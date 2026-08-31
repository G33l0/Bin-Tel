"""Excel rendering for Bin-Tel reports.

Produces a workbook with a branded cover sheet carrying the report's identity,
summary and criteria, and a separate records sheet with a frozen, filterable
header — the shape a spreadsheet user actually wants to receive.
"""

from __future__ import annotations

import io
from typing import TYPE_CHECKING, Any

from app.core.constants import APP_NAME, APP_TAGLINE, APP_VERSION, COPYRIGHT
from app.core.logging_config import get_logger

if TYPE_CHECKING:  # pragma: no cover - typing only
    from app.services.report_service import ReportContent

logger = get_logger(__name__)

INK = "FF101A2B"
TEAL = "FF158F86"
BAND = "FFF2F6FB"
RULE = "FFD8E0EC"
MUTED = "FF5B6B82"

MAX_COLUMN_WIDTH = 46


def render_xlsx(content: ReportContent) -> bytes:
    """Render *content* to an .xlsx workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    workbook.properties.creator = f"{APP_NAME} {APP_VERSION}"
    workbook.properties.title = content.title
    workbook.properties.subject = content.report_type.label

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    title_font = Font(name="Calibri", size=18, bold=True, color=INK)
    subtitle_font = Font(name="Calibri", size=11, color=MUTED)
    section_font = Font(name="Calibri", size=11, bold=True, color=TEAL)
    label_font = Font(name="Calibri", size=10, bold=True)
    body_font = Font(name="Calibri", size=10)
    header_fill = PatternFill("solid", fgColor=INK)
    band_fill = PatternFill("solid", fgColor=BAND)
    thin = Side(style="thin", color=RULE)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # -- cover sheet -------------------------------------------------------
    cover = workbook.active
    cover.title = "Report"
    cover.sheet_view.showGridLines = False
    cover.column_dimensions["A"].width = 30
    cover.column_dimensions["B"].width = 68

    row = 1
    cover.cell(row=row, column=1, value=APP_NAME).font = title_font
    cover.cell(row=row, column=2, value=APP_TAGLINE).font = subtitle_font
    row += 2

    cover.cell(row=row, column=1, value=content.title).font = Font(
        name="Calibri", size=14, bold=True, color=INK
    )
    row += 1
    if content.subtitle:
        cover.cell(row=row, column=1, value=content.subtitle).font = subtitle_font
        row += 1
    row += 1

    def write_block(heading: str, rows: list[tuple[str, str]]) -> None:
        nonlocal row
        if not rows:
            return
        cover.cell(row=row, column=1, value=heading).font = section_font
        row += 1
        for label, value in rows:
            cover.cell(row=row, column=1, value=str(label)).font = label_font
            cell = cover.cell(row=row, column=2, value=_coerce(value))
            cell.font = body_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        row += 1

    write_block(
        "Report details",
        [
            ("Generated", content.generated_at.strftime("%d %B %Y, %H:%M UTC")),
            ("Database version", content.database_version or "Unknown"),
            ("Report type", content.report_type.label),
            ("Records", f"{content.row_count:,}"),
        ],
    )
    write_block("Search criteria", content.criteria)
    write_block("Summary", content.summary)
    for heading, rows in content.detail_sections:
        write_block(heading, rows)

    for note in content.notes:
        cover.cell(row=row, column=1, value=note).font = subtitle_font
        row += 1
    row += 1
    cover.cell(row=row, column=1, value=COPYRIGHT).font = subtitle_font

    # -- records sheet -----------------------------------------------------
    if content.table_columns and content.table_rows:
        sheet = workbook.create_sheet("Records")
        sheet.sheet_view.showGridLines = False

        for index, column in enumerate(content.table_columns, start=1):
            cell = sheet.cell(row=1, column=index, value=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")

        for row_index, values in enumerate(content.table_rows, start=2):
            for column_index, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_index, column=column_index, value=_coerce(value))
                cell.font = body_font
                cell.border = border
                if row_index % 2 == 0:
                    cell.fill = band_fill

        # A frozen, filterable header is what makes a large export usable.
        sheet.freeze_panes = "A2"
        last_column = get_column_letter(len(content.table_columns))
        sheet.auto_filter.ref = f"A1:{last_column}{len(content.table_rows) + 1}"
        sheet.row_dimensions[1].height = 20

        for index, column in enumerate(content.table_columns, start=1):
            widest = max(
                [len(str(column))]
                + [
                    len(str(row[index - 1]))
                    for row in content.table_rows[:500]
                    if index - 1 < len(row)
                ]
            )
            sheet.column_dimensions[get_column_letter(index)].width = min(
                MAX_COLUMN_WIDTH, max(10, widest + 2)
            )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _coerce(value: Any) -> Any:
    """Write numbers as numbers so Excel can sum and sort them."""
    text = str(value).strip()
    if not text or text in ("—", "Unknown"):
        return text
    cleaned = text.replace(",", "")
    if cleaned.isdigit() and len(cleaned) <= 15 and not text.startswith("0"):
        # A BIN is an identifier, not a quantity; leading zeros must survive,
        # and a long digit run stays text so Excel cannot reformat it.
        if len(cleaned) >= 6:
            return text
        return int(cleaned)
    try:
        if cleaned.count(".") == 1 and cleaned.replace(".", "").isdigit():
            return float(cleaned)
    except ValueError:  # pragma: no cover - defensive
        pass
    return text
